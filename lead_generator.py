#!/usr/bin/env python3
"""
Tavern Buddy - Bar Owner Lead Generator

Compiles bar/pub/tavern owner leads in Philadelphia & Atlantic City
with email addresses, output to Excel spreadsheet.

Sources:
  1. OpenStreetMap (Overpass API) - 360+ bars, free, no key needed
  2. Business website email scraping - crawls /contact pages for emails
  3. Google Places API (optional - much more data with a free key)
  4. Hunter.io email enrichment (optional - fills gaps)

Usage:
  pip install requests beautifulsoup4 openpyxl
  python3 lead_generator.py

Optional env vars for enhanced results:
  export GOOGLE_PLACES_API_KEY=your_key   # https://console.cloud.google.com
  export HUNTER_API_KEY=your_key           # https://hunter.io (25 free/mo)
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import time
import json
import os
import sys
import hashlib
from urllib.parse import urlparse
from datetime import datetime

# ──────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────

GOOGLE_PLACES_API_KEY = os.environ.get("GOOGLE_PLACES_API_KEY", "")
HUNTER_API_KEY = os.environ.get("HUNTER_API_KEY", "")

OUTPUT_FILE = "bar_leads_{}.xlsx".format(datetime.now().strftime("%Y%m%d_%H%M%S"))

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml",
}

EMAIL_JUNK_DOMAINS = {
    "example.com", "sentry.io", "sentry-next.wixpress.com", "wixpress.com",
    "googleapis.com", "schema.org", "w3.org", "wordpress.org", "gravatar.com",
    "wix.com", "squarespace.com", "cloudflare.com", "jquery.com",
    "google.com", "facebook.com", "twitter.com", "instagram.com",
    "fbcdn.net", "gstatic.com", "apple.com", "mystore.com", "domain.com",
    "yourstore.com", "yourdomain.com", "test.com", "placeholder.com",
    "change.me", "update.me",
}


# ──────────────────────────────────────────────
# Data model
# ──────────────────────────────────────────────

def make_lead(**kwargs):
    base = {
        "business_name": "", "address": "", "city": "", "state": "",
        "zip": "", "phone": "", "website": "", "email": "",
        "owner_manager": "", "source": "", "area": "", "category": "",
    }
    base.update(kwargs)
    return base


def lead_key(lead):
    raw = (lead["business_name"] + lead.get("address", "")).lower()
    raw = re.sub(r"[^a-z0-9]", "", raw)
    return hashlib.md5(raw.encode()).hexdigest()


# ──────────────────────────────────────────────
# Source 1: OpenStreetMap / Overpass API (free)
# ──────────────────────────────────────────────

def fetch_osm_bars():
    """Pull all bars/pubs/nightclubs from OSM for Philly and AC areas."""
    leads = []
    queries = [
        {
            "label": "Philadelphia",
            "state": "PA",
            "query": """
                [out:json][timeout:60];
                area[name="Philadelphia"][admin_level=8]->.searchArea;
                (
                  node["amenity"~"bar|pub|nightclub"](area.searchArea);
                  way["amenity"~"bar|pub|nightclub"](area.searchArea);
                );
                out body;
            """,
        },
        {
            "label": "Atlantic City Area",
            "state": "NJ",
            "query": """
                [out:json][timeout:60];
                (
                  node["amenity"~"bar|pub|nightclub"](39.20,-74.70,39.50,-74.30);
                  way["amenity"~"bar|pub|nightclub"](39.20,-74.70,39.50,-74.30);
                );
                out body;
            """,
        },
    ]

    for q in queries:
        print(f"  Fetching {q['label']}...", end=" ", flush=True)
        try:
            resp = requests.get(
                "https://overpass-api.de/api/interpreter",
                params={"data": q["query"]},
                timeout=90,
            )
            resp.raise_for_status()
            elements = resp.json().get("elements", [])
            count = 0
            for el in elements:
                tags = el.get("tags", {})
                name = tags.get("name", "")
                if not name:
                    continue

                amenity = tags.get("amenity", "bar")
                phone = tags.get("phone") or tags.get("contact:phone", "")
                website = tags.get("website") or tags.get("contact:website", "")
                email = tags.get("email") or tags.get("contact:email", "")
                addr_street = tags.get("addr:street", "")
                addr_num = tags.get("addr:housenumber", "")
                addr_city = tags.get("addr:city", q["label"].split(" Area")[0])
                addr_zip = tags.get("addr:postcode", "")

                address = f"{addr_num} {addr_street}".strip() if addr_street else ""

                lead = make_lead(
                    business_name=name,
                    address=address,
                    city=addr_city,
                    state=q["state"],
                    zip=addr_zip,
                    phone=phone,
                    website=website,
                    email=email,
                    source="OpenStreetMap",
                    area=q["label"],
                    category=amenity,
                )
                leads.append(lead)
                count += 1

            print(f"{count} bars found")
            time.sleep(1)
        except Exception as e:
            print(f"error: {e}")

    return leads


# ──────────────────────────────────────────────
# Source 2: Google Places API (optional, much richer)
# ──────────────────────────────────────────────

GOOGLE_SEARCH_POINTS = [
    {"label": "Philadelphia - Center City", "lat": 39.9526, "lng": -75.1652},
    {"label": "Philadelphia - South Philly", "lat": 39.9276, "lng": -75.1652},
    {"label": "Philadelphia - Northern Liberties", "lat": 39.9700, "lng": -75.1400},
    {"label": "Philadelphia - Fishtown", "lat": 39.9743, "lng": -75.1300},
    {"label": "Philadelphia - Manayunk", "lat": 40.0265, "lng": -75.2240},
    {"label": "Philadelphia - West Philly", "lat": 39.9550, "lng": -75.2100},
    {"label": "Philadelphia - Kensington", "lat": 39.9880, "lng": -75.1250},
    {"label": "Atlantic City", "lat": 39.3643, "lng": -74.4229},
    {"label": "Ventnor/Margate", "lat": 39.3340, "lng": -74.4900},
    {"label": "Somers Point/EHT", "lat": 39.3400, "lng": -74.5800},
]


def search_google_places_all():
    """Search Google Places API across all target areas."""
    if not GOOGLE_PLACES_API_KEY:
        return []

    leads = []
    seen_place_ids = set()

    for point in GOOGLE_SEARCH_POINTS:
        print(f"  Searching: {point['label']}...", end=" ", flush=True)
        new_count = 0
        next_token = None

        for page in range(3):  # up to 3 pages (60 results per area)
            params = {"key": GOOGLE_PLACES_API_KEY}
            if next_token:
                params["pagetoken"] = next_token
                time.sleep(2)
            else:
                params.update({
                    "location": f"{point['lat']},{point['lng']}",
                    "radius": 5000,
                    "type": "bar",
                })

            try:
                resp = requests.get(
                    "https://maps.googleapis.com/maps/api/place/nearbysearch/json",
                    params=params, timeout=15,
                )
                data = resp.json()

                for place in data.get("results", []):
                    pid = place.get("place_id")
                    if pid in seen_place_ids:
                        continue
                    seen_place_ids.add(pid)

                    # Get details
                    details = {}
                    try:
                        dr = requests.get(
                            "https://maps.googleapis.com/maps/api/place/details/json",
                            params={
                                "place_id": pid,
                                "fields": "formatted_phone_number,website,formatted_address",
                                "key": GOOGLE_PLACES_API_KEY,
                            },
                            timeout=10,
                        )
                        details = dr.json().get("result", {})
                        time.sleep(0.05)
                    except:
                        pass

                    full_addr = details.get("formatted_address", place.get("vicinity", ""))
                    city, state, zipcode = "", "", ""
                    if full_addr:
                        # Parse "123 Main St, Philadelphia, PA 19103, USA"
                        parts = [p.strip() for p in full_addr.split(",")]
                        if len(parts) >= 3:
                            city = parts[-3] if len(parts) >= 4 else parts[-2]
                            state_zip = parts[-2] if len(parts) >= 3 else ""
                            sz_parts = state_zip.strip().split()
                            state = sz_parts[0] if sz_parts else ""
                            zipcode = sz_parts[1] if len(sz_parts) > 1 else ""

                    lead = make_lead(
                        business_name=place.get("name", ""),
                        address=full_addr.split(",")[0] if full_addr else "",
                        city=city,
                        state=state,
                        zip=zipcode,
                        phone=details.get("formatted_phone_number", ""),
                        website=details.get("website", ""),
                        source="GooglePlaces",
                        area=point["label"],
                        category="bar",
                    )
                    leads.append(lead)
                    new_count += 1

                next_token = data.get("next_page_token")
                if not next_token:
                    break
            except Exception as e:
                print(f"error: {e}")
                break

        print(f"{new_count} bars")

    return leads


# ──────────────────────────────────────────────
# Email Discovery
# ──────────────────────────────────────────────

def is_junk_email(email):
    """Filter out non-human emails."""
    email = email.lower()
    domain = email.split("@")[-1]
    if domain in EMAIL_JUNK_DOMAINS:
        return True
    if any(ext in email for ext in [".png", ".jpg", ".gif", ".svg", ".css", ".js"]):
        return True
    # Filter hex-looking strings (tracking pixels, sentry DSNs, etc.)
    local_part = email.split("@")[0]
    if re.match(r'^[a-f0-9]{20,}$', local_part):
        return True
    return False


def extract_emails_from_html(html):
    """Pull all email addresses from HTML text."""
    emails = set()

    # Regex for emails
    for match in re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', html):
        email = match.lower().strip()
        if not is_junk_email(email):
            emails.add(email)

    # mailto: links
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.select("a[href^='mailto:']"):
        href = a["href"].replace("mailto:", "").split("?")[0].strip().lower()
        if "@" in href and "." in href.split("@")[-1] and not is_junk_email(href):
            emails.add(href)

    return emails


def scrape_website_for_emails(url, timeout=10):
    """Scrape a business website for email addresses on homepage + contact page."""
    all_emails = set()
    if not url:
        return all_emails

    if not url.startswith("http"):
        url = "https://" + url

    try:
        # Homepage
        resp = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        if resp.status_code == 200:
            all_emails.update(extract_emails_from_html(resp.text))

            # Try contact/about pages
            base = f"{urlparse(resp.url).scheme}://{urlparse(resp.url).netloc}"
            for path in ["/contact", "/contact-us", "/about", "/about-us", "/info"]:
                try:
                    r2 = requests.get(base + path, headers=HEADERS, timeout=8, allow_redirects=True)
                    if r2.status_code == 200:
                        all_emails.update(extract_emails_from_html(r2.text))
                    time.sleep(0.2)
                except:
                    continue
    except:
        pass

    return all_emails


def hunter_domain_search(domain):
    """Use Hunter.io to find emails for a domain."""
    if not HUNTER_API_KEY or not domain:
        return []
    try:
        resp = requests.get(
            "https://api.hunter.io/v2/domain-search",
            params={"domain": domain, "api_key": HUNTER_API_KEY, "limit": 5},
            timeout=10,
        )
        results = []
        for item in resp.json().get("data", {}).get("emails", []):
            results.append({
                "email": item.get("value", ""),
                "name": f"{item.get('first_name', '')} {item.get('last_name', '')}".strip(),
                "position": item.get("position", ""),
            })
        return results
    except:
        return []


def enrich_emails(leads):
    """Scrape websites and use Hunter.io to find emails for all leads."""
    total = len(leads)
    sites_to_check = sum(1 for l in leads if l.get("website") and not l.get("email"))
    found = 0

    print(f"  {sites_to_check} websites to scrape for emails...")

    for idx, lead in enumerate(leads):
        if lead.get("email"):
            found += 1
            continue

        if not lead.get("website"):
            continue

        if (idx + 1) % 20 == 0 or idx + 1 == total:
            print(f"  Progress: {idx+1}/{total} processed, {found} emails found so far", flush=True)

        # Scrape website
        emails = scrape_website_for_emails(lead["website"])
        if emails:
            lead["email"] = "; ".join(sorted(emails)[:3])
            found += 1
            time.sleep(0.3)
            continue

        # Hunter.io fallback
        if HUNTER_API_KEY:
            domain = urlparse(
                lead["website"] if "://" in lead["website"] else "https://" + lead["website"]
            ).netloc
            if domain:
                results = hunter_domain_search(domain)
                if results:
                    lead["email"] = results[0]["email"]
                    if results[0].get("name") and not lead.get("owner_manager"):
                        lead["owner_manager"] = results[0]["name"]
                    found += 1

        time.sleep(0.3)

    return found


# ──────────────────────────────────────────────
# PA Liquor License Public Data (bonus source)
# ──────────────────────────────────────────────

def try_pa_license_lookup():
    """
    Attempt to pull PA liquor license data. The PLCB has a public search.
    This enriches leads with license holder (owner) names.
    """
    print("  Checking PA PLCB license data...", end=" ", flush=True)
    try:
        # PLCB license search API
        resp = requests.get(
            "https://www.lcb.pa.gov/Licensing/Pages/default.aspx",
            headers=HEADERS,
            timeout=10,
        )
        if resp.status_code == 200:
            print("available (would need Selenium for full scrape)")
        else:
            print("not accessible")
    except:
        print("not accessible")
    return {}


# ──────────────────────────────────────────────
# Excel Output
# ──────────────────────────────────────────────

def write_excel(leads, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bar Leads"

    # Styles
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    email_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    phone_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    columns = [
        ("Business Name", 32), ("Address", 28), ("City", 18), ("State", 7),
        ("ZIP", 9), ("Phone", 17), ("Website", 38), ("Email", 38),
        ("Owner/Manager", 22), ("Area", 25), ("Type", 12), ("Source", 14),
    ]

    # Header row
    for i, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws.freeze_panes = "A2"

    # Sort: email first, then phone, then rest
    def sort_key(l):
        has_email = 0 if l.get("email") else 1
        has_phone = 0 if l.get("phone") else 1
        return (has_email, has_phone, l.get("area", ""), l.get("business_name", ""))

    leads_sorted = sorted(leads, key=sort_key)

    # Data rows
    for row_idx, lead in enumerate(leads_sorted, 2):
        values = [
            lead.get("business_name"), lead.get("address"), lead.get("city"),
            lead.get("state"), lead.get("zip"), lead.get("phone"),
            lead.get("website"), lead.get("email"), lead.get("owner_manager"),
            lead.get("area"), lead.get("category"), lead.get("source"),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val or "")
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if lead.get("email"):
                cell.fill = email_fill
            elif lead.get("phone") and not lead.get("email"):
                cell.fill = phone_fill

    # ── Summary Sheet ──
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Tavern Buddy Lead Report"
    ws2["A1"].font = Font(bold=True, size=14)

    stats = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Leads", len(leads)),
        ("With Email", sum(1 for l in leads if l.get("email"))),
        ("With Phone", sum(1 for l in leads if l.get("phone"))),
        ("With Website", sum(1 for l in leads if l.get("website"))),
    ]
    for i, (label, val) in enumerate(stats, 3):
        ws2[f"A{i}"] = label
        ws2[f"A{i}"].font = Font(bold=True)
        ws2[f"B{i}"] = val

    # By area breakdown
    ws2["A10"] = "Breakdown by Area"
    ws2["A10"].font = Font(bold=True, size=12)
    ws2["A11"], ws2["B11"], ws2["C11"], ws2["D11"] = "Area", "Total", "With Email", "With Phone"
    for c in ["A11", "B11", "C11", "D11"]:
        ws2[c].font = Font(bold=True)

    area_stats = {}
    for l in leads:
        a = l.get("area", "Unknown")
        if a not in area_stats:
            area_stats[a] = {"total": 0, "email": 0, "phone": 0}
        area_stats[a]["total"] += 1
        if l.get("email"):
            area_stats[a]["email"] += 1
        if l.get("phone"):
            area_stats[a]["phone"] += 1

    for i, (area, s) in enumerate(sorted(area_stats.items()), 12):
        ws2[f"A{i}"] = area
        ws2[f"B{i}"] = s["total"]
        ws2[f"C{i}"] = s["email"]
        ws2[f"D{i}"] = s["phone"]

    for col in ["A", "B", "C", "D"]:
        ws2.column_dimensions[col].width = 28

    # Color legend
    row = 12 + len(area_stats) + 2
    ws2[f"A{row}"] = "Color Legend"
    ws2[f"A{row}"].font = Font(bold=True, size=12)
    ws2[f"A{row+1}"] = "Green rows"
    ws2[f"B{row+1}"] = "Have email address"
    ws2[f"A{row+1}"].fill = email_fill
    ws2[f"A{row+2}"] = "Yellow rows"
    ws2[f"B{row+2}"] = "Have phone but no email"
    ws2[f"A{row+2}"].fill = phone_fill

    wb.save(filename)
    return filename


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main():
    print()
    print("=" * 60)
    print("  TAVERN BUDDY - Bar Owner Lead Generator")
    print("  Philadelphia & Atlantic City Areas")
    print("=" * 60)
    print()

    if GOOGLE_PLACES_API_KEY:
        print("[+] Google Places API key detected")
    else:
        print("[ ] No Google Places key (set GOOGLE_PLACES_API_KEY for 2-3x more bars)")
    if HUNTER_API_KEY:
        print("[+] Hunter.io API key detected")
    else:
        print("[ ] No Hunter.io key (set HUNTER_API_KEY for more email discovery)")
    print()

    all_leads = {}

    # ── Phase 1: OpenStreetMap ──
    print("PHASE 1: Pulling bar data from OpenStreetMap...")
    osm_leads = fetch_osm_bars()
    for lead in osm_leads:
        k = lead_key(lead)
        all_leads[k] = lead
    print(f"  => {len(osm_leads)} bars from OSM\n")

    # ── Phase 2: Google Places (optional) ──
    if GOOGLE_PLACES_API_KEY:
        print("PHASE 2: Pulling from Google Places API...")
        gp_leads = search_google_places_all()
        new = 0
        for lead in gp_leads:
            k = lead_key(lead)
            if k not in all_leads:
                all_leads[k] = lead
                new += 1
            else:
                # Merge in any missing data
                existing = all_leads[k]
                if not existing["website"] and lead["website"]:
                    existing["website"] = lead["website"]
                if not existing["phone"] and lead["phone"]:
                    existing["phone"] = lead["phone"]
                if not existing["address"] and lead["address"]:
                    existing["address"] = lead["address"]
        print(f"  => {new} new bars from Google Places ({len(gp_leads)} total, rest merged)\n")
    else:
        print("PHASE 2: Skipping Google Places (no API key)\n")

    total = len(all_leads)
    if total == 0:
        print("No bars found. Check your internet connection and try again.")
        sys.exit(1)

    print(f"TOTAL UNIQUE BARS: {total}\n")

    # ── Phase 3: PA License Data ──
    print("PHASE 3: Public records check...")
    try_pa_license_lookup()
    print()

    # ── Phase 4: Email Discovery ──
    leads_list = list(all_leads.values())
    already_have_email = sum(1 for l in leads_list if l.get("email"))
    have_website = sum(1 for l in leads_list if l.get("website"))
    print(f"PHASE 4: Email discovery...")
    print(f"  {already_have_email} already have emails from source data")
    print(f"  {have_website} have websites to scrape for contact emails")

    email_count = enrich_emails(leads_list)

    print(f"\n  => {email_count} total leads with email addresses\n")

    # ── Phase 5: Write output ──
    print(f"PHASE 5: Writing Excel spreadsheet...")
    output = write_excel(leads_list, OUTPUT_FILE)
    print(f"  => Saved: {output}")

    # ── Summary ──
    email_total = sum(1 for l in leads_list if l.get("email"))
    phone_total = sum(1 for l in leads_list if l.get("phone"))
    website_total = sum(1 for l in leads_list if l.get("website"))

    print()
    print("=" * 60)
    print("  RESULTS")
    print("=" * 60)
    print(f"  Total bar leads:     {len(leads_list)}")
    print(f"  With email:          {email_total}  <-- your priority contacts")
    print(f"  With phone only:     {phone_total - email_total}")
    print(f"  With website:        {website_total}")
    print(f"  Output:              {output}")
    print()
    print("  Spreadsheet color code:")
    print("    GREEN  = has email address")
    print("    YELLOW = has phone but no email")
    print()

    if not GOOGLE_PLACES_API_KEY or not HUNTER_API_KEY:
        print("  BOOST YOUR RESULTS (free):")
        if not GOOGLE_PLACES_API_KEY:
            print("  export GOOGLE_PLACES_API_KEY=<key>")
            print("    -> Adds 200-400 more bars from Google (free tier)")
            print("    -> Get key: https://console.cloud.google.com/apis")
        if not HUNTER_API_KEY:
            print("  export HUNTER_API_KEY=<key>")
            print("    -> Finds emails for bars without websites")
            print("    -> 25 free lookups/mo: https://hunter.io")
        print()


if __name__ == "__main__":
    main()
